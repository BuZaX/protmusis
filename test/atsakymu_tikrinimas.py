import openpyxl
from openpyxl import load_workbook



def write_score(link, players, score_list, stage):
    no_of_questions = 5
    wb = load_workbook(filename=link + 'rezultatai.xlsx')
    ws = wb.active
    print(score_list)

    for i in range(0, len(score_list)):
        for j in range(0, len(score_list[i])):
            ws.cell(i+2, stage*(no_of_questions+2)+j-4, int(score_list[i][j]))

    wb.save(filename=link + 'rezultatai.xlsx')

def read_player(lnk, players, stage, correct):
    score_list = []
    for player in players:
        ats_arr = []
        score = []
        with open(lnk+player+"_"+str(stage)+".txt", "r") as ats:
            for cnt, line in enumerate(ats):
                ats_arr.append(line.strip())
                if correct[stage-1][cnt][0][0] == ats_arr[cnt]:
                    print("Question " + str(cnt+1) + ": Award " + str(
                        correct[stage-1][cnt][0][2]) + " points to player " + player)
                    score.append(correct[stage-1][cnt][0][2])
                else:
                    print("Question " + str(cnt + 1) + ": Award 0 points to player " + player)
                    score.append(0)
        score_list.append(score)
    return score_list


def read_correct(lnk):
    with open(lnk, "r") as ats:
        ats = open(lnk, "r")
        ats_arr = []
        sec_arr = []
        for cnt, line in enumerate(ats):
            if line[0] != '-':
                ats_arr.append([line.strip()])
            elif line[0] == '-':
                sec_arr.append(ats_arr)
                ats_arr = []
    return sec_arr


def main():
    players = ["dominykas", "petras", "jonas", "Inga", "Dovydas"]
    link = "C:/Users/dominykas.gudavicius/Desktop/test/"

    stage = 1

    correct_answer = read_correct(link+"atsakymai.txt")
    score = read_player(link, players, stage, correct_answer)
    write_score(link, players, score, stage)
if __name__ == '__main__':
    main()